import openpyxl
from django.core.management.base import BaseCommand
from django.conf import settings
from django.db import connection
from loans.models import Customer, Loan
from dateutil import parser as date_parser
import os


class Command(BaseCommand):
    help = 'Ingest customer and loan data from Excel files into the database'

    def handle(self, *args, **kwargs):
        # Skip if data already loaded
        if Customer.objects.exists():
            self.stdout.write(self.style.WARNING('Data already ingested. Skipping.'))
            return

        self.stdout.write('Starting data ingestion...')
        self.ingest_customers()
        self.ingest_loans()
        self.stdout.write(self.style.SUCCESS('Data ingestion complete!'))

    def ingest_customers(self):
        filepath = settings.DATA_DIR / 'customer_data.xlsx'
        if not os.path.exists(filepath):
            self.stdout.write(self.style.ERROR(f'File not found: {filepath}'))
            return

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        customers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Excel columns: Customer ID, First Name, Last Name, Age, Phone Number, Monthly Salary, Approved Limit
            customer_id, first_name, last_name, age, phone_number, monthly_salary, approved_limit = row

            if customer_id is None:
                continue

            customers.append(Customer(
                customer_id=int(customer_id),
                first_name=str(first_name) if first_name else '',
                last_name=str(last_name) if last_name else '',
                age=int(age) if age else None,
                phone_number=int(phone_number) if phone_number else 0,
                monthly_salary=int(monthly_salary) if monthly_salary else 0,
                approved_limit=int(approved_limit) if approved_limit else 0,
                current_debt=0.0,
            ))

        Customer.objects.bulk_create(customers, ignore_conflicts=True)
        self.stdout.write(self.style.SUCCESS(f'  Ingested {len(customers)} customers'))

        # Reset auto-increment sequence so new registrations don't clash with ingested IDs
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT setval(
                    pg_get_serial_sequence('loans_customer', 'customer_id'),
                    (SELECT MAX(customer_id) FROM loans_customer)
                )
            """)
        self.stdout.write(self.style.SUCCESS('  Reset customer ID sequence'))

    def ingest_loans(self):
        filepath = settings.DATA_DIR / 'loan_data.xlsx'
        if not os.path.exists(filepath):
            self.stdout.write(self.style.ERROR(f'File not found: {filepath}'))
            return

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        loans = []
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Excel columns: Customer ID, Loan ID, Loan Amount, Tenure, Interest Rate, Monthly Payment, EMIs Paid on Time, Date of Approval, End Date
            customer_id, loan_id, loan_amount, tenure, interest_rate, monthly_repayment, emis_paid_on_time, start_date, end_date = row

            if loan_id is None or customer_id is None:
                continue

            try:
                customer = Customer.objects.get(customer_id=int(customer_id))
            except Customer.DoesNotExist:
                skipped += 1
                continue

            # Parse dates safely
            parsed_start = self._parse_date(start_date)
            parsed_end = self._parse_date(end_date)

            loans.append(Loan(
                loan_id=int(loan_id),
                customer=customer,
                loan_amount=float(loan_amount) if loan_amount else 0.0,
                tenure=int(tenure) if tenure else 0,
                interest_rate=float(interest_rate) if interest_rate else 0.0,
                monthly_repayment=float(monthly_repayment) if monthly_repayment else 0.0,
                emis_paid_on_time=int(emis_paid_on_time) if emis_paid_on_time else 0,
                start_date=parsed_start,
                end_date=parsed_end,
            ))

        Loan.objects.bulk_create(loans, ignore_conflicts=True)
        self.stdout.write(self.style.SUCCESS(f'  Ingested {len(loans)} loans (skipped {skipped})'))

        # Reset auto-increment sequence so new loans don't clash with ingested IDs
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT setval(
                    pg_get_serial_sequence('loans_loan', 'loan_id'),
                    (SELECT MAX(loan_id) FROM loans_loan)
                )
            """)
        self.stdout.write(self.style.SUCCESS('  Reset loan ID sequence'))

    def _parse_date(self, value):
        if value is None:
            return None
        if hasattr(value, 'date'):
            return value.date()
        try:
            return date_parser.parse(str(value)).date()
        except Exception:
            return None